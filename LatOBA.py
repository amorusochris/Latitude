#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import pandas as pd
import xlsxwriter
import openpyxl
import datetime

#Server File location
loc2 = ('S:\OPERATIONS & ENGINEERING\Well DR Pressure\Dashboard - Wells.xlsm')
#Local file location for off network
#loc2 =('C:\\Users\camoruso\Desktop\Dashboard - Wells.xlsm')

sheetname = "WS Alt Data"

#login Info from local file
F1 = open("C:\\Users\camoruso\Desktop\password.txt", "r")
contents = F1.readlines()
username = contents[0]
password = contents[1]
F1.close()

#Open Worksheet with Pandas
wellsdf = pd.read_excel(loc2, sheet_name = sheetname)
Excelrowcount = len(wellsdf.index)
columns = wellsdf.columns
column = columns[0]
Datelist = wellsdf[column].tolist()
last = len(Datelist)
lastdate = Datelist[last-1]

today = datetime.date.today()
Yesterday = datetime.date.today() + datetime.timedelta(days=-1)
if lastdate.date() < Yesterday:
	Begdate = lastdate.date() + datetime.timedelta(days=1)
	Begdate = Begdate.strftime('%m/%d/%Y')
	print(Begdate)
	Enddate = Yesterday
	Enddate = Enddate.strftime('%m/%d/%Y')
	print(Enddate)
else:
	Begdate = "05/02/2019"
	Enddate= "05/07/2019"

driver = webdriver.Chrome()
#"C:\Users\camoruso\Documents\Helpful Docs\Programming")
driver.get('https://www.gasstorage.net/WORSHAMSTEED/Operator/index.cfm')

element = driver.find_element_by_xpath('//*[@id="username"]')
element.send_keys(username)

element = driver.find_element_by_xpath('//*[@id="password"]')
element.send_keys(password)

element = driver.find_element_by_xpath('//*[@id="frmLogin"]/table[1]/tbody/tr[3]/td/input')
element.send_keys(Keys.RETURN)

element = driver.find_element_by_xpath('//*[@id="navmenu"]/table/tbody/tr[14]/td/a')
element.send_keys(Keys.RETURN)

element = driver.find_element_by_xpath('//*[@id="NetFacilStart_Disp"]')
element.clear()
element.send_keys(Begdate)

element = driver.find_element_by_xpath('//*[@id="NetFacilEnd_Disp"]')
element.clear()
element.send_keys(Enddate)

element = driver.find_element_by_xpath('//*[@id="content"]/div/div[3]/fieldset[3]/form/table/tbody/tr/td[5]/input')
element.send_keys(Keys.RETURN)

table = driver.find_element_by_xpath('//*[@id="content"]/div/div[3]/table/tbody')
#table2 = driver.find_elements_by_xpath('//*[@id="content"]/div/div[3]/table/tbody/tr')
rows = table.find_elements_by_tag_name("tr")

x= []
for row in rows:
	x.append([td.text for td in row.find_elements_by_tag_name("td")])

OBAdf = pd.DataFrame(x)

#Close webdriver
driver.close()
driver.quit()

#Count number of rows copied over
obadatarows = len(OBAdf.index)
#print(df[1:numrows2-1])

#Change dataframe to chop off header and total row
OBAdf2 = OBAdf[1:obadatarows-1]
print(OBAdf2.dtypes)

#for col in OBAdf2.columns[2:]:
#	OBAdf2[col] = OBAdf2[col].apply(pd.to_numeric)
	#convert_object(convert_numeric = True)
#print(OBAdf2.dtypes)
	
#output
#print(df[1:numrows2])
#writer = pd.Excelwriter(loc2)
#Write to excel after the number of entries from origina numrows (Excel sheet count)
#with pd.ExcelWriter(loc2) as writer:
#	OBAdf2.to_excel(writer,sheetname,startrow=Excelrowcount+2)
#	writer.save	

outputfile = ('S:\OPERATIONS & ENGINEERING\Well DR Pressure\OBA.xlsx')
OBAdf2.to_excel(outputfile)
print(OBAdf2.values.tolist())

wb = openpyxl.load_workbook(loc2,keep_vba=True)
ws = wb.get_sheet_by_name(sheetname)
wb2 = openpyxl.load_workbook(outputfile)
ws2 = wb2.worksheets[0]

for i in range(2,obadatarows):
	for j in range(1,6):
			ws.cell(Excelrowcount+i,j).value = (ws2.cell(i,j+1).value)
			print(ws2.cell(i,j+1).value)

wb.save(loc2)
